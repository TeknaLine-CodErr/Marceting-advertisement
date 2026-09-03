# -*- coding: utf-8 -*-
"""Выгрузка справочного отчёта в Excel с условным форматированием по статусу."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

STATUS_FILL = {
    "green": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "yellow": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "red": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "no_co": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    "no_price_data": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    "error": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}

COLUMNS = [
    ("article", "Артикул", 22),
    ("name", "Название", 45),
    ("category", "Категория", 28),
    ("scheme", "Схема", 10),
    ("source_file", "Файл ЦО", 30),
    ("price_co", "Наша цена, ₽", 12),
    ("price_promo_max", "Макс. цена акции, ₽", 16),
    ("price_used", "Цена участия, ₽", 14),
    ("pct_of_cost", "% от себестоимости", 16),
    ("net_profit", "Чистая прибыль, ₽", 16),
    ("status_label", "Статус", 24),
    ("price_change_note", "Изменить цену в ЦО", 22),
    ("reason", "Комментарий", 55),
    ("ordered_qty_7d", "Заказано, шт/нед", 14),
    ("stock_ozon", "Остаток Ozon, шт", 14),
    ("stock_own", "Остаток на своём складе, шт", 16),
    ("stock_note", "Рекомендация по остаткам", 45),
]


def export_report(report_rows, promo_name, thresholds, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт по акции"

    title = f"Отчёт по товарам для участия в акции" + (f' «{promo_name}»' if promo_name else "")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = (
        f"Пороги одобрения по % от себестоимости: зелёный >= {thresholds.green_min:.0%}, "
        f"жёлтый {thresholds.red_max:.0%}-{thresholds.green_min:.0%}, красный < {thresholds.red_max:.0%}. "
        f"Рекомендация по остаткам (справочно, не влияет на статус): много остатков — от "
        f"{thresholds.stock_days_high:.0f} дн. запаса, низкие продажи — до "
        f"{thresholds.low_sales_per_week:.0f} шт/нед"
    )
    ws["A2"].font = Font(italic=True, size=9)

    header_row = 4
    for i, (_, label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=i, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    for r, row in enumerate(report_rows, start=header_row + 1):
        fill = STATUS_FILL.get(row.get("status"))
        for c, (field, _, _) in enumerate(COLUMNS, start=1):
            value = row.get(field)
            cell = ws.cell(row=r, column=c, value=value)
            if field == "pct_of_cost" and isinstance(value, (int, float)):
                cell.number_format = "0.0%"
            elif field in ("price_co", "price_promo_max", "price_used", "net_profit", "ordered_qty_7d", "stock_ozon", "stock_own") and isinstance(value, (int, float)):
                cell.number_format = "#,##0"
            if fill:
                cell.fill = fill
            if field in ("reason", "stock_note"):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            if field == "price_change_note" and row.get("price_change_needed"):
                cell.font = Font(bold=True, color="9C5700")
            if field == "stock_note" and row.get("stock_recommendation") in ("high_stock", "low_sales"):
                cell.font = Font(italic=True, color="1F6FB2")

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(COLUMNS))}{header_row + len(report_rows)}"

    wb.save(out_path)
    return out_path

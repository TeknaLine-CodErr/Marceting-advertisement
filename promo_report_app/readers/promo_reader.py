# -*- coding: utf-8 -*-
"""Чтение файла акции Ozon (шаблон вида "Максимальный бустинг.xlsx")."""

import os
import openpyxl

from config import (
    PROMO_COLUMNS,
    PROMO_DATA_START_ROW,
    PROMO_DESCRIPTION_SHEET,
)


def _find_main_sheet(wb):
    for name in wb.sheetnames:
        if name != PROMO_DESCRIPTION_SHEET:
            return name
    return None


def load_promo_file(path):
    """Возвращает (catalog: {article: row_dict}, promo_name: str, warnings: [str])"""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    warnings = []
    fname = os.path.basename(path)

    promo_name = None
    if PROMO_DESCRIPTION_SHEET in wb.sheetnames:
        desc_ws = wb[PROMO_DESCRIPTION_SHEET]
        first_line = desc_ws["A1"].value
        if first_line:
            promo_name = str(first_line).split("—")[0].replace("Название акции:", "").strip()

    sheet_name = _find_main_sheet(wb)
    if sheet_name is None:
        warnings.append(f"В файле {fname} не найден лист с товарами акции")
        wb.close()
        return {}, promo_name, warnings

    ws = wb[sheet_name]
    catalog = {}
    for r in range(PROMO_DATA_START_ROW, ws.max_row + 1):
        article = ws[f"{PROMO_COLUMNS['article']}{r}"].value
        if article in (None, ""):
            continue
        row = {}
        for field, col in PROMO_COLUMNS.items():
            row[field] = ws[f"{col}{r}"].value
        if article in catalog:
            warnings.append(f"Артикул «{article}» дублируется в файле акции {fname} (строка {r} проигнорирована)")
            continue
        catalog[article] = row

    wb.close()
    return catalog, promo_name, warnings

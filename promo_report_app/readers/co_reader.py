# -*- coding: utf-8 -*-
"""Чтение файлов "Ценообразование" (.xlsm): листы СП / FBS / Неликвид + справочник Категории."""

import os
import openpyxl

from config import (
    CO_COLUMNS,
    CO_DATA_START_ROW,
    CO_ROW1_CONSTS,
    CO_SCHEMES,
    CO_NON_SCHEME_SHEETS,
    CATEGORY_SHEET_NAME,
    CATEGORY_MATCH_COLUMN,
    CATEGORY_BRACKET_LOW_ROW,
    CATEGORY_BRACKET_HIGH_ROW,
    CATEGORY_DATA_START_ROW,
)


def list_available_schemes(path):
    """Возвращает список названий листов из CO_SCHEMES, реально присутствующих в файле
    (в порядке CO_SCHEMES), чтобы UI показывал чекбоксы только для существующих листов."""
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        sheet_names = set(wb.sheetnames)
    finally:
        wb.close()
    return [s for s in CO_SCHEMES if s in sheet_names and s not in CO_NON_SCHEME_SHEETS]


def load_category_rates(wb):
    """Возвращает {тип_товара: {"FBO": [(low, high, rate), ...], "FBS": [...]}}"""
    if CATEGORY_SHEET_NAME not in wb.sheetnames:
        return {}
    ws = wb[CATEGORY_SHEET_NAME]

    col_sets = {"FBO": ("C", "D", "E"), "FBS": ("F", "G", "H")}
    bounds = {}
    for label, cols in col_sets.items():
        triples = []
        for col in cols:
            low = ws[f"{col}{CATEGORY_BRACKET_LOW_ROW}"].value
            high = ws[f"{col}{CATEGORY_BRACKET_HIGH_ROW}"].value
            triples.append((col, low, high))
        bounds[label] = triples

    result = {}
    r = CATEGORY_DATA_START_ROW
    empty_streak = 0
    while empty_streak < 50:
        type_name = ws[f"{CATEGORY_MATCH_COLUMN}{r}"].value
        if type_name in (None, ""):
            empty_streak += 1
            r += 1
            continue
        empty_streak = 0
        entry = {}
        for label, colbounds in bounds.items():
            rates = []
            for col, low, high in colbounds:
                rate = ws[f"{col}{r}"].value
                rates.append((low, high, rate))
            entry[label] = rates
        result[type_name] = entry
        r += 1
    return result


def _read_row1_consts(ws):
    return {name: ws[cell_ref].value for name, cell_ref in CO_ROW1_CONSTS.items()}


def _detect_commission_label(rows, category_rates):
    """Определяет, какой набор колонок тарифной сетки ("FBO" колонки C:E или "FBS" F:H
    в листе Категории) в реальности использует этот лист схемы — сравнивая уже
    посчитанную в ЦО ставку комиссии (кэш) с обеими сетками при текущей цене товара.
    Работает по названию листа-схемы независимо (Китай/Россия/Монеты и т.п.)."""
    scores = {"FBO": 0, "FBS": 0}
    for row in rows:
        rates = category_rates.get(row["category"])
        if not rates:
            continue
        cached_rate = row.get("commission_rate")
        price = row.get("price")
        if not isinstance(cached_rate, (int, float)) or not isinstance(price, (int, float)):
            continue
        for label in ("FBO", "FBS"):
            for low, high, rate in rates.get(label) or []:
                if low is None or high is None or rate is None:
                    continue
                if low <= price <= high and abs(rate - cached_rate) < 1e-6:
                    scores[label] += 1
                    break
    if scores["FBO"] == 0 and scores["FBS"] == 0:
        return None
    return "FBO" if scores["FBO"] >= scores["FBS"] else "FBS"


def load_co_file(path, schemes_to_include):
    """Читает выбранные листы одного файла ЦО.

    Возвращает (catalog: {article: row_dict}, warnings: [str])
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    category_rates = load_category_rates(wb)

    catalog = {}
    warnings = []
    fname = os.path.basename(path)

    for scheme in schemes_to_include:
        if scheme not in wb.sheetnames:
            warnings.append(f"Лист «{scheme}» не найден в файле {fname}")
            continue
        ws = wb[scheme]
        row1 = _read_row1_consts(ws)

        sheet_rows = []
        for r in range(CO_DATA_START_ROW, ws.max_row + 1):
            article = ws[f"{CO_COLUMNS['article']}{r}"].value
            if article in (None, ""):
                continue
            in_sale = ws[f"{CO_COLUMNS['in_sale']}{r}"].value
            if in_sale != "Да":
                continue
            price = ws[f"{CO_COLUMNS['price']}{r}"].value
            if not isinstance(price, (int, float)) or price <= 0:
                continue

            row = {}
            for field, col in CO_COLUMNS.items():
                row[field] = ws[f"{col}{r}"].value
            row["article"] = article
            row["scheme"] = scheme
            row["source_file"] = fname
            row["row1_consts"] = row1
            row["category_rates"] = category_rates.get(row["category"])
            sheet_rows.append(row)

        commission_label = _detect_commission_label(sheet_rows, category_rates)
        if commission_label is None and sheet_rows:
            warnings.append(
                f"Лист «{scheme}» в файле {fname}: не удалось определить тарифную сетку "
                f"комиссии (FBO/FBS) — комиссия при новой цене будет приблизительной "
                f"(используется текущая ставка из ЦО)"
            )

        for row in sheet_rows:
            row["commission_label"] = commission_label
            article = row["article"]
            if article in catalog:
                warnings.append(
                    f"Артикул «{article}» встречается повторно "
                    f"(лист {scheme}, файл {fname}) — использована первая найденная запись"
                )
                continue
            catalog[article] = row

    wb.close()
    return catalog, warnings


def load_co_files(file_scheme_pairs):
    """file_scheme_pairs: [(path, [scheme, ...]), ...] до 6 файлов.

    Возвращает (catalog: {article: row_dict}, warnings: [str]) — объединённый каталог.
    """
    merged = {}
    all_warnings = []
    for path, schemes in file_scheme_pairs:
        if not schemes:
            continue
        catalog, warnings = load_co_file(path, schemes)
        all_warnings.extend(warnings)
        for article, row in catalog.items():
            if article in merged:
                all_warnings.append(
                    f"Артикул «{article}» уже был загружен из другого файла/листа "
                    f"({merged[article]['source_file']}, {merged[article]['scheme']}) — "
                    f"запись из {row['source_file']} ({row['scheme']}) проигнорирована"
                )
                continue
            merged[article] = row
    return merged, all_warnings
